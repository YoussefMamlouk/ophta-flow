import pandas as pd
from typing import List, Dict, Optional
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
import logging

logger = logging.getLogger(__name__)


def merge_excel_data(
    existing_excel_path: Optional[str],
    new_rows: List[Dict[str, any]],
    output_path: str
) -> str:
    """
    Merge new rows into existing Excel file, preserving all formatting, colors, and formulas.
    Adds data to the "donnée" sheet only.
    
    Args:
        existing_excel_path: Path to existing Excel file (None if no existing file)
        new_rows: List of dictionaries representing new rows to add
        output_path: Path where output Excel file will be saved
        
    Returns:
        Path to the output Excel file
    """
    logger.info(f"merge_excel_data called with {len(new_rows)} new rows")
    logger.info(f"Existing Excel path: {existing_excel_path}")
    logger.info(f"Output path: {output_path}")
    
    if not existing_excel_path:
        # If no existing file, create a new one
        logger.info("No existing Excel file, creating new one")
        return create_excel_from_rows(new_rows, output_path)
    
    # Load the workbook preserving all formatting
    # Don't use keep_vba unless necessary as it can cause corruption
    logger.info(f"Loading workbook from: {existing_excel_path}")
    wb = load_workbook(existing_excel_path, data_only=False)
    logger.info(f"Workbook loaded. Sheets: {wb.sheetnames}")
    
    # Find or create the "donnée" sheet
    sheet_name = "donnée"
    if sheet_name not in wb.sheetnames:
        # If sheet doesn't exist, create it
        logger.warning(f"Sheet '{sheet_name}' not found, creating it")
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
        logger.info(f"Found sheet '{sheet_name}'")
    
    # Find column headers in row 2 (1-indexed, so row 2 = index 1)
    header_row = 2
    column_map = {}  # Maps column name to column index (1-indexed)
    
    logger.info(f"Looking for headers in row {header_row}. Sheet max_column: {ws.max_column}, max_row: {ws.max_row}")
    
    # Read headers from row 2
    # Check up to a reasonable number of columns (e.g., 100)
    max_cols_to_check = max(ws.max_column, 100) if ws.max_column > 0 else 100
    logger.info(f"Checking up to {max_cols_to_check} columns")
    
    for col_idx in range(1, max_cols_to_check + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            header_value = str(cell.value).strip()
            if header_value:  # Only add non-empty headers
                column_map[header_value] = col_idx
                logger.debug(f"Found header '{header_value}' at column {col_idx}")
    
    if not column_map:
        # If no headers found, try row 1
        logger.info("No headers found in row 2, trying row 1")
        header_row = 1
        for col_idx in range(1, max_cols_to_check + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.value:
                header_value = str(cell.value).strip()
                if header_value:  # Only add non-empty headers
                    column_map[header_value] = col_idx
                    logger.debug(f"Found header '{header_value}' at column {col_idx}")
    
    logger.info(f"Found {len(column_map)} column headers: {list(column_map.keys())}")
    
    # Find the first available (empty) row after the header
    # Start from header_row + 1 (first data row)
    first_available_row = header_row + 1
    
    # Check all rows after header to find the first empty row
    # An empty row is one where key columns (ID Patient and Œil) are empty
    id_patient_col = column_map.get("ID Patient")
    oeil_col = column_map.get("Œil")
    
    # First, find the last row with actual data (for duplicate checking)
    last_data_row = header_row
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Check if row has any data in mapped columns
        has_data = False
        for col_idx in column_map.values():
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip():
                has_data = True
                break
        if has_data:
            last_data_row = row_idx
    
    # Find first empty row (where ID Patient and Œil are both empty)
    # Start checking from header_row + 1
    if id_patient_col and oeil_col:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            id_cell = ws.cell(row=row_idx, column=id_patient_col)
            eye_cell = ws.cell(row=row_idx, column=oeil_col)
            id_value = str(id_cell.value or "").strip()
            eye_value = str(eye_cell.value or "").strip()
            # If both ID and Eye are empty, this is the first available row
            if not id_value and not eye_value:
                first_available_row = row_idx
                break
        else:
            # No empty row found, use the row after the last data row
            first_available_row = last_data_row + 1 if last_data_row > header_row else header_row + 1
    else:
        # Fallback: use row after last data row
        first_available_row = last_data_row + 1 if last_data_row > header_row else header_row + 1
    
    logger.info(f"First available row: {first_available_row}, Last data row: {last_data_row}")
    
    # Get existing patient ID + eye combinations for duplicate detection
    existing_keys = set()
    
    if id_patient_col and oeil_col:
        # Check all rows from header+1 to max_row for existing data
        for row_idx in range(header_row + 1, ws.max_row + 1):
            patient_id = str(ws.cell(row=row_idx, column=id_patient_col).value or "").strip()
            eye = str(ws.cell(row=row_idx, column=oeil_col).value or "").strip()
            if patient_id and eye:
                existing_keys.add((patient_id, eye))
    
    # Filter out duplicates from new rows
    filtered_rows = []
    logger.info(f"Processing {len(new_rows)} new rows. Existing keys: {len(existing_keys)}")
    for idx, row in enumerate(new_rows):
        patient_id = str(row.get("ID Patient", "")).strip()
        eye = str(row.get("Œil", "")).strip()
        logger.debug(f"Row {idx}: ID={patient_id}, Eye={eye}, Row data keys: {list(row.keys())}")
        if patient_id and eye and (patient_id, eye) not in existing_keys:
            filtered_rows.append(row)
            existing_keys.add((patient_id, eye))  # Prevent duplicates within new_rows
            logger.debug(f"Row {idx} added (not duplicate)")
        else:
            logger.debug(f"Row {idx} skipped: patient_id={patient_id}, eye={eye}, is_duplicate={(patient_id, eye) in existing_keys}")
    
    logger.info(f"After filtering: {len(filtered_rows)} rows to add")
    
    # Add new rows
    if filtered_rows:
        logger.info(f"Starting to add {len(filtered_rows)} rows. First available row: {first_available_row}, Last data row: {last_data_row}, Header row: {header_row}")
        # Get a sample row to copy formatting from
        # Prefer last data row, otherwise use header row (for formatting reference)
        sample_row = None
        if last_data_row > header_row:
            sample_row = last_data_row
            logger.info(f"Using last data row {sample_row} as formatting template")
        elif ws.max_row >= header_row:
            # Use header row as formatting template if no data rows exist
            sample_row = header_row
            logger.info(f"Using header row {sample_row} as formatting template")
        else:
            logger.warning(f"No sample row available. max_row={ws.max_row}, header_row={header_row}")
        
        # Start writing from the first available row
        current_row = first_available_row
        for row_idx, row_data in enumerate(filtered_rows):
            new_row_idx = current_row
            logger.info(f"Processing row {row_idx + 1}/{len(filtered_rows)}: Adding to row {new_row_idx}")
            logger.info(f"  Row data keys: {list(row_data.keys())}")
            logger.info(f"  Row data values: {row_data}")
            
            # Copy row formatting if sample row exists
            if sample_row and sample_row > 0 and sample_row <= ws.max_row:
                # Copy row height
                if sample_row in ws.row_dimensions and ws.row_dimensions[sample_row].height:
                    ws.row_dimensions[new_row_idx].height = ws.row_dimensions[sample_row].height
                
                # Copy cell formatting from sample row
                # Only copy formatting for columns that exist in the sheet
                max_col_to_copy = max(
                    ws.max_column if ws.max_column > 0 else 0,
                    max(column_map.values()) if column_map else 0,
                    100  # Reasonable upper limit
                )
                for col_idx in range(1, max_col_to_copy + 1):
                    source_cell = ws.cell(row=sample_row, column=col_idx)
                    target_cell = ws.cell(row=new_row_idx, column=col_idx)
                    
                    # Copy cell formatting safely
                    try:
                        if source_cell.has_style:
                            if source_cell.font:
                                target_cell.font = copy(source_cell.font)
                            if source_cell.border:
                                target_cell.border = copy(source_cell.border)
                            if source_cell.fill:
                                target_cell.fill = copy(source_cell.fill)
                            if source_cell.number_format:
                                target_cell.number_format = copy(source_cell.number_format)
                            if source_cell.protection:
                                target_cell.protection = copy(source_cell.protection)
                            if source_cell.alignment:
                                target_cell.alignment = copy(source_cell.alignment)
                    except Exception:
                        # If copying formatting fails, continue without it
                        pass
            
            # Write data to appropriate columns only
            # This preserves any formulas in other columns
            logger.info(f"Writing data to row {new_row_idx}")
            written_count = 0
            for field_name, value in row_data.items():
                if field_name in column_map:
                    col_idx = column_map[field_name]
                    cell = ws.cell(row=new_row_idx, column=col_idx)
                    
                    # Convert value to appropriate type
                    converted_value = value
                    if value is not None and value != "":
                        # Try to convert numeric strings to numbers
                        if isinstance(value, str):
                            # Clean Patient ID - extract only digits and convert to integer
                            if field_name == "ID Patient":
                                import re
                                digits_only = re.sub(r'\D', '', value)
                                if digits_only:
                                    try:
                                        converted_value = int(digits_only)
                                    except ValueError:
                                        converted_value = digits_only
                                else:
                                    converted_value = value.strip()
                            # Convert "Modèle implanté" to integer (it's a number)
                            elif field_name == "Modèle implanté":
                                try:
                                    # Extract digits and convert to integer
                                    import re
                                    digits_only = re.sub(r'\D', '', value)
                                    if digits_only:
                                        converted_value = int(digits_only)
                                    else:
                                        converted_value = value.strip()
                                except (ValueError, AttributeError):
                                    converted_value = value.strip()
                            # Try to convert to number for numeric fields
                            elif field_name in ["AL", "ACD epit", "LT", "PACHY (mm)", "WTW (mm)", 
                                               "PUISSANCE IOL", "TORIQUE PROG EDOF", "K1", "K2", "Âge"]:
                                try:
                                    # Replace comma with dot for decimal separator
                                    num_str = value.replace(",", ".").strip()
                                    if "." in num_str:
                                        converted_value = float(num_str)
                                    else:
                                        converted_value = int(num_str)
                                except (ValueError, AttributeError):
                                    # If conversion fails, keep as string
                                    converted_value = value
                    
                    # Set the value (cell is new/empty, so no formula to preserve)
                    # Formatting already copied above
                    cell.value = converted_value
                    written_count += 1
                    logger.info(f"  ✓ Written '{field_name}' = '{converted_value}' ({type(converted_value).__name__}) to column {col_idx} (row {new_row_idx})")
                else:
                    logger.warning(f"  ✗ Field '{field_name}' not found in column_map")
            
            logger.info(f"Row {new_row_idx}: Written {written_count}/{len(row_data)} fields")
            # Move to next row for next iteration
            current_row += 1
    else:
        logger.warning("No filtered rows to add!")
    
    # Save the workbook with proper error handling
    logger.info(f"Saving workbook to {output_path}")
    try:
        wb.save(output_path)
        logger.info("Workbook saved successfully")
    except Exception as e:
        logger.error(f"Failed to save workbook: {str(e)}")
        raise Exception(f"Failed to save Excel file: {str(e)}")
    finally:
        # Always close the workbook
        try:
            wb.close()
            logger.info("Workbook closed")
        except Exception:
            pass  # Ignore errors on close
    
    return output_path


def create_excel_from_rows(new_rows: List[Dict[str, any]], output_path: str) -> str:
    """
    Create a new Excel file from rows (when no existing Excel provided).
    Creates a "donnée" sheet with proper structure.
    
    Args:
        new_rows: List of dictionaries representing rows
        output_path: Path where output Excel file will be saved
        
    Returns:
        Path to the output Excel file
    """
    logger.info(f"create_excel_from_rows called with {len(new_rows)} rows")
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "donnée"
    logger.info(f"Created new workbook with sheet '{ws.title}'")
    
    # Define expected columns
    columns = [
        "ID Patient", "Âge", "Œil", "AL", "PACHY (mm)", "ACD epit",
        "LT", "PUISSANCE IOL", "TORIQUE PROG EDOF", "WTW (mm)"
    ]
    
    # Write empty row 1
    for col_idx, _ in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value="")
    
    # Write headers in row 2
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=2, column=col_idx, value=col_name)
    
    # Write data starting from row 3
    if new_rows:
        logger.info(f"Writing {len(new_rows)} rows starting from row 3")
        for row_idx, row_data in enumerate(new_rows, start=3):
            logger.debug(f"Writing row {row_idx}: {row_data}")
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
                logger.debug(f"  Written '{col_name}' = '{value}' to row {row_idx}, column {col_idx}")
    else:
        logger.warning("No rows to write")
    
    # Save and close the workbook
    logger.info(f"Saving new workbook to {output_path}")
    try:
        wb.save(output_path)
        logger.info("New workbook saved successfully")
    except Exception as e:
        logger.error(f"Failed to save new workbook: {str(e)}")
        raise Exception(f"Failed to save Excel file: {str(e)}")
    finally:
        wb.close()
        logger.info("New workbook closed")
    
    return output_path

